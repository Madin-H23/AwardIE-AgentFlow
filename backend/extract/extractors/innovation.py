"""
大创项目抽取器

从Excel文件中抽取大学生创新创业训练项目数据。

适配新框架 backend/extract/framework.py
"""
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from backend.extract.extractors.base import Extractor, ExtractContext
from backend.extract.types import (
    ExtractResult, ExtractStatus, TemplateType,
    ValidationResult, ValidationError
)
from backend.extract.exceptions import ExtractorError

logger = logging.getLogger(__name__)

# 尝试导入依赖库
try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("需要安装 openpyxl: pip install openpyxl")

try:
    import xlrd
except ImportError:
    xlrd = None  # xlrd 是可选的


class ColumnMapper:
    """列名映射器"""

    def __init__(self, column_mapping: Dict[str, List[str]]):
        """
        初始化列名映射器

        Args:
            column_mapping: 配置中的列名映射
        """
        self.mapping = column_mapping or {}

    @staticmethod
    def normalize(col_name: str) -> str:
        """清洗列名：去除空格、回车等"""
        if col_name is None:
            return ""
        return str(col_name).strip().replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")

    def find_column(self, columns: List[str], internal_field: str) -> Optional[str]:
        """
        在列名列表中查找对应内部字段的列名

        Args:
            columns: 列名列表
            internal_field: 内部标准字段名

        Returns:
            找到的实际列名，未找到返回None
        """
        possible_names = self.mapping.get(internal_field, [])
        columns_normalized = {self.normalize(col): col for col in columns}

        for possible in possible_names:
            normalized = self.normalize(possible)
            if normalized in columns_normalized:
                return columns_normalized[normalized]
        return None

    def get_value(self, row_dict: Dict[str, str], internal_field: str) -> Optional[str]:
        """
        从行数据中获取指定字段的值

        Args:
            row_dict: 行数据字典（列名 -> 值）
            internal_field: 内部标准字段名

        Returns:
            字段值（去除前后空格），如果未找到返回None
        """
        possible_names = self.mapping.get(internal_field, [])

        for col_name in possible_names:
            normalized_col = self.normalize(col_name)
            for row_col, row_value in row_dict.items():
                if self.normalize(row_col) == normalized_col:
                    value = row_value.strip()
                    return value if value else None

        return None


class DataParser:
    """数据解析器"""

    # 学号长度固定9位
    STUDENT_ID_PATTERN = re.compile(r'^\d{9}$')
    # 姓名模式：2-5字符，不包含数字
    NAME_PATTERN = re.compile(r'^[^\d]{2,5}$')

    @classmethod
    def parse_teachers(cls, teachers_str: Any) -> List[str]:
        """
        解析指导教师字符串

        支持多种分隔符：逗号、顿号、分号、空格等
        """
        if teachers_str is None or not str(teachers_str).strip():
            return []

        text = str(teachers_str)
        # 统一分隔符为逗号
        text = text.replace('、', ',').replace('，', ',').replace(';', ',').replace('；', ',')
        text = text.replace('\u3000', ',').replace('\t', ',')

        return [t.strip() for t in text.split(',') if t.strip()]

    @classmethod
    def parse_students(cls, members_str: Any) -> List[Dict[str, str]]:
        """
        解析学生成员

        支持多种格式：
        - 姓名/学号：张三/212203227
        - 姓名（学号）：张三（212203227）
        - 学号+姓名：212203227张三
        - 姓名直接连接学号：张三212203227
        """
        if members_str is None or not str(members_str).strip():
            return []

        text = str(members_str)
        # 替换中文标点为英文标点
        text = text.replace('，', ',').replace('、', ',').replace('；', ';')
        text = text.replace('\n', ' ').replace('\r', ' ')
        text = text.strip()

        # 先尝试按逗号或分号分隔
        if ',' in text or ';' in text:
            parts = re.split(r'[,;]+', text)
            students = []
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                student = cls._parse_single_student(part)
                if student:
                    students.append(student)
            return students

        # 如果没有逗号或分号，使用正则表达式方法
        text = re.sub(r'[;]+', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()

        students = []
        seen_ids = {}

        # 多种正则模式（T49/T75 挂账修复：每项附 (姓名组号, 学号组号)——
        # 原 P3「学号+姓名」分组序反排，旧循环统一按 g1=姓名 消费致该格式整体失效）
        patterns = [
            (r'([^\d\s\(\)]+?)\s*/\s*(\d{9})', 1, 2),         # 姓名/学号
            (r'([^\d\s\(\)]+?)\s*[\(（]\s*(\d{9})\s*[\)）]', 1, 2),  # 姓名（学号）
            (r'(\d{9})\s*([^\d\s（）()]*)', 2, 1),             # 学号+姓名
            (r'([^\d\s（）()]+?)(\d{9})', 1, 2),               # 姓名直接连接学号（排除括号防吞全角括号）
        ]

        for pattern, name_g, id_g in patterns:
            for match in re.finditer(pattern, text):
                name = match.group(name_g).strip()
                student_id = match.group(id_g).strip()
                name = name.rstrip('/,，;、；').strip()

                # 验证姓名格式
                if cls.NAME_PATTERN.match(name):
                    if student_id not in seen_ids or len(name) > len(seen_ids[student_id][0]):
                        seen_ids[student_id] = (name, student_id)

        for name, student_id in seen_ids.values():
            students.append({"姓名": name, "学号": student_id})

        return students

    @classmethod
    def _parse_single_student(cls, text: str) -> Optional[Dict[str, str]]:
        """解析单个学生字符串"""
        if not text or not text.strip():
            return None

        text = text.strip()
        re.sub(r'\s+', ' ', text)

        # 多种正则模式（学号固定9位）
        patterns = [
            (r'([^\d\s\(\)]+?)\s+(\d{9})', 1, 2),           # 姓名 学号
            (r'([^\d\s\(\)]+?)\s*/\s*(\d{9})', 1, 2),        # 姓名/学号
            (r'([^\d\s\(\)]+?)\s*[\(（]\s*(\d{9})\s*[\)）]', 1, 2),  # 姓名（学号）
            (r'(\d{9})\s*([^\d\s\(\)]+)', 1, 2),             # 学号+姓名
            (r'([^\d\s（）()]+?)(\d{9})', 1, 2),              # 姓名直接连接学号（同上：排除括号）
        ]

        for pattern, name_group, id_group in patterns:
            match = re.match(pattern, text)
            if match:
                name = match.group(name_group).strip()
                student_id = match.group(id_group).strip()

                # 验证姓名和学号格式
                if cls.NAME_PATTERN.match(name) and cls.STUDENT_ID_PATTERN.match(student_id):
                    return {"姓名": name, "学号": student_id}

        return None

    @classmethod
    def parse_date_range(cls, date_str: Any) -> tuple:
        """
        解析起讫时间

        格式如：2024.6-2025.6 或 2024.01-2025.12
        """
        if date_str is None or not str(date_str).strip():
            return None, None

        text = str(date_str)
        text = text.replace('～', '-').replace('—', '-').replace('~', '-')

        match = re.search(r'(\d{4})\.?(\d{1,2})?\s*[-－—～]\s*(\d{4})\.?(\d{1,2})?', text)
        if match:
            start_year = match.group(1)
            start_month = match.group(2) or '01'
            end_year = match.group(3)
            end_month = match.group(4) or '12'
            return (
                f"{start_year}.{start_month.zfill(2)}",
                f"{end_year}.{end_month.zfill(2)}"
            )

        return None, None

    @classmethod
    def parse_year_to_date_range(cls, year_value: Any) -> tuple:
        """
        将年份转换为日期范围

        立项文件中只有年份，假设从当年6月到次年6月
        """
        if year_value is None:
            return None, None

        year_str = str(int(year_value)) if isinstance(year_value, (int, float)) else str(year_value)
        if not year_str.isdigit():
            return None, None

        return f"{year_str}.06", f"{int(year_str) + 1}.06"

    @classmethod
    def validate_student_id(cls, student_id: str) -> bool:
        """验证学号格式（固定9位数字）"""
        return bool(cls.STUDENT_ID_PATTERN.match(student_id))

    @classmethod
    def validate_name(cls, name: str) -> bool:
        """验证姓名格式（2-5字符，不包含数字）"""
        return bool(cls.NAME_PATTERN.match(name))


class InnovationExtractor(Extractor):
    """
    大创项目抽取器

    从Excel文件中抽取大学生创新创业训练项目数据。
    """

    name = "innovation"
    description = "大学生创新创业训练项目文件"
    judgment_text = "通常包含：项目名称 + 项目负责人 + 项目成员 + 指导教师 + 起讫时间 + 项目级别 + 年份"
    keywords = []  # 不使用关键词匹配，基于文件扩展名

    def __init__(self, config: Dict[str, Any]):
        """
        初始化抽取器

        Args:
            config: 从配置加载的配置字典
        """
        self._config = config
        self._extensions = config.get("extensions", [".xlsx", ".xls"])
        self._target_departments = config.get("target_departments", ["计算机工程系"])
        self._header_keywords = config.get("header_keywords", ["项目", "学号", "教师", "负责人"])
        self._file_keywords = config.get("file_keywords", [])
        self._column_mapper = ColumnMapper(config.get("column_mapping", {}))
        self._student_id_length = config.get("student_id_length", 9)

        logger.info(f"大创抽取器初始化: 目标系别={self._target_departments}")

    @property
    def extensions(self) -> List[str]:
        """支持的文件扩展名"""
        return self._extensions

    def matches_extension(self, ext: str) -> bool:
        """检查文件扩展名是否支持"""
        return ext.lower() in [e.lower() for e in self._extensions]

    def matches_keywords(self, text: str) -> bool:
        """大创抽取器不使用关键词匹配（基于文件扩展名）"""
        return False

    def extract(self, ctx: ExtractContext) -> ExtractResult:
        """
        执行抽取

        Args:
            ctx: 抽取上下文

        Returns:
            抽取结果
        """
        file_path = ctx.file_path

        try:
            # 1. 读取Excel文件
            rows = self._read_excel(file_path)
            if not rows:
                return self._other_result("无法读取Excel文件")

            filename = Path(file_path).name

            # 2. 检查是否为大创文件
            if not self._is_innovation_file(rows, filename):
                return self._other_result("不是大创文件")

            # 3. 查找表头行
            header_row_idx = self._find_header_row(rows)
            if header_row_idx is None:
                return self._other_result("未找到有效的表头行")

            # 4. 提取列名
            columns = rows[header_row_idx]

            # 5. 从第一行提取默认级别和年份
            first_row_str = ' '.join(rows[0]) if rows else ''
            default_level = self._extract_level_from_first_row(first_row_str)
            default_year = self._extract_year_from_first_row(first_row_str)

            logger.info(f"从第一行提取的默认值: 级别={default_level}, 年份={default_year}")

            # 6. 找到含有"系"的列用于筛选
            dept_col_idx = self._find_department_column(columns)

            # 7. 筛选数据行
            data_rows = rows[header_row_idx + 1:]

            if dept_col_idx is not None:
                data_rows = [
                    row for row in data_rows
                    if len(row) > dept_col_idx
                    and self._contains_department(row[dept_col_idx])
                ]

            if not data_rows:
                return self._other_result("没有相关数据")

            # 8. 抽取项目
            projects = self._extract_projects(data_rows, columns, default_level, default_year)

            # 9. 验证数据并生成 ValidationResult
            validated, validation_result = self._validate(projects)

            logger.info(f"大创抽取成功: {len(validated)} 个项目")

            return ExtractResult(
                status=ExtractStatus.SUCCESS,
                data={
                    "projects": validated,
                    "count": len(validated)
                },
                template_type=TemplateType.INNOVATION,
                extractor_name=self.name,
                validation_result=validation_result,
                metadata={
                    "source_file": file_path,
                    "header_row": header_row_idx,
                    "filtered_rows": len(data_rows),
                    "target_department": self._target_departments[0] if self._target_departments else None
                }
            )

        except ExtractorError as e:
            logger.error(f"大创抽取失败: {e}")
            return ExtractResult(
                status=ExtractStatus.PARSE_ERROR,
                error_message=str(e),
                template_type=TemplateType.OTHER,
                extractor_name=self.name,
            )
        except Exception as e:
            logger.exception(f"大创抽取异常: {e}")
            return ExtractResult(
                status=ExtractStatus.FILE_ERROR,
                error_message=f"处理失败: {e}",
                template_type=TemplateType.OTHER,
                extractor_name=self.name,
            )

    def _read_excel(self, file_path: str) -> List[List]:
        """读取Excel文件"""
        path = Path(file_path)

        if path.suffix.lower() == '.xlsx':
            return self._read_xlsx(file_path)
        elif path.suffix.lower() == '.xls':
            return self._read_xls(file_path)
        else:
            raise ExtractorError(f"不支持的文件格式: {path.suffix}")

    def _read_xlsx(self, file_path: str) -> List[List]:
        """读取 .xlsx 文件"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            wb.close()
            return rows
        except Exception as e:
            raise ExtractorError(f"读取 .xlsx 文件失败: {e}")

    def _read_xls(self, file_path: str) -> List[List]:
        """读取 .xls 文件"""
        if xlrd is None:
            raise ExtractorError("xlrd 未安装，无法读取 .xls 文件")

        try:
            workbook = xlrd.open_workbook(file_path)
            worksheet = workbook.sheet_by_index(0)
            rows = []
            for row_idx in range(worksheet.nrows):
                row = worksheet.row(row_idx)
                rows.append([str(cell.value) if cell.value else "" for cell in row])
            return rows
        except Exception as e:
            raise ExtractorError(f"读取 .xls 文件失败: {e}")

    def _is_innovation_file(self, rows: List[List], filename: str) -> bool:
        """检查是否为大创文件"""
        # 检查第一行
        if rows and len(rows) > 0:
            first_row_str = ' '.join(rows[0])
            for keyword in self._file_keywords:
                if keyword in first_row_str:
                    logger.debug(f"通过第一行识别为大创文件 (关键词: {keyword})")
                    return True

        # 检查文件名
        for keyword in self._file_keywords:
            if keyword in filename:
                logger.debug(f"通过文件名识别为大创文件 (关键词: {keyword})")
                return True

        return False

    def _find_header_row(self, rows: List[List]) -> Optional[int]:
        """查找表头行"""
        for idx, row in enumerate(rows):
            row_str = ' '.join(row)
            keyword_count = sum(1 for kw in self._header_keywords if kw in row_str)
            if keyword_count >= 2:
                logger.debug(f"找到表头行: {idx}, 关键词数: {keyword_count}")
                return idx

        return None

    def _find_department_column(self, columns: List[str]) -> Optional[int]:
        """找到含有"系"的列"""
        for idx, col in enumerate(columns):
            if '系' in col:
                logger.debug(f"找到系别列: {idx} - '{col}'")
                return idx
        return None

    def _contains_department(self, cell_value: str) -> bool:
        """检查单元格值是否包含目标系别"""
        if not cell_value:
            return False
        return any(dept in cell_value for dept in self._target_departments)

    def _extract_level_from_first_row(self, first_row_str: str) -> str:
        """从第一行提取项目级别"""
        if '国家级' in first_row_str:
            return '国家级'
        elif '省级' in first_row_str:
            return '省级'
        elif '院级' in first_row_str:
            return '院级'
        return ''

    def _extract_year_from_first_row(self, first_row_str: str) -> Optional[int]:
        """从第一行提取年份"""
        match = re.search(r'(\d{4})年', first_row_str)
        if match:
            return int(match.group(1))
        return None

    def _extract_projects(
        self,
        rows: List[List],
        columns: List[str],
        default_level: str = '',
        default_year: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """抽取项目数据"""
        projects = []

        for row in rows:
            # 跳过空行
            if not any(cell.strip() for cell in row):
                continue

            # 构建行字典
            row_dict = {}
            for i, cell_value in enumerate(row):
                if i < len(columns):
                    row_dict[columns[i]] = cell_value

            # 尝试抽取项目
            project = self._extract_single_row(row_dict, default_level, default_year)
            if project:
                projects.append(project)

        return projects

    def _extract_single_row(
        self,
        row_dict: Dict[str, str],
        default_level: str = '',
        default_year: Optional[int] = None
    ) -> Optional[Dict[str, Any]]:
        """从单行抽取项目数据"""
        # 必须有项目名称
        project_name = self._column_mapper.get_value(row_dict, "项目名称")
        if not project_name:
            return None

        # 解析学号/负责人
        leader_id = self._column_mapper.get_value(row_dict, "项目负责人学号")
        leader_name = self._column_mapper.get_value(row_dict, "学生负责人")

        student_leader = None
        if leader_name and leader_id:
            # 验证学号
            if DataParser.validate_student_id(leader_id):
                student_leader = {"姓名": leader_name, "学号": leader_id}
        elif leader_name:
            student_leader = {"姓名": leader_name, "学号": ""}

        # 解析其他成员
        members_str = self._column_mapper.get_value(row_dict, "项目其他成员信息")
        other_members = []
        if members_str:
            other_members = DataParser.parse_students(members_str)

        # 解析指导教师
        teachers_str = self._column_mapper.get_value(row_dict, "指导教师")
        teachers = DataParser.parse_teachers(teachers_str) if teachers_str else []

        # 解析时间字段 - 直接从列映射获取可能的列名
        date_value = None
        for col_name in ["起讫时间", "项目开始时间", "起止时间", "起止年月", "立项年份"]:
            date_value = self._column_mapper.get_value(row_dict, col_name)
            if date_value:
                break

        start_date = None
        end_date = None
        if date_value:
            # 处理字符串格式的日期范围
            date_str = str(date_value).strip()
            if '.' in date_str and '-' in date_str:
                start_date, end_date = DataParser.parse_date_range(date_str)
            elif date_str.isdigit() and len(date_str) == 4:
                start_date, end_date = DataParser.parse_year_to_date_range(date_str)

        # 提取年份
        year = None
        if start_date:
            year_match = re.match(r'^(\d{4})\.', start_date)
            if year_match:
                year = int(year_match.group(1))

        if year is None and default_year is not None:
            year = default_year

        # 提取项目级别
        level = self._column_mapper.get_value(row_dict, "项目级别")
        if not level and default_level:
            level = default_level

        # 格式化日期：将 2025.05 改为 2025-05
        def format_date(date_str):
            if date_str and '.' in date_str:
                return date_str.replace('.', '-')
            return date_str

        # 格式化成员：将字典数组改为字符串数组 ["姓名(学号)"]
        def format_members(members_list):
            if not members_list:
                return []
            return [f"{m['姓名']}({m['学号']})" for m in members_list]

        # 从Excel中读取系别，如果没有则使用配置中的目标系别
        department = self._column_mapper.get_value(row_dict, "系别")
        if not department:
            department = self._target_departments[0] if self._target_departments else ""

        return {
            "project_number": self._column_mapper.get_value(row_dict, "项目编号"),
            "project_name": project_name,
            "start_date": format_date(start_date),
            "end_date": format_date(end_date),
            "leader_name": student_leader["姓名"] if student_leader else None,
            "leader_student_id": student_leader["学号"] if student_leader else None,
            "members": format_members(other_members),
            "supervisors": teachers,
            "project_level": level,
            "acceptance_level": self._column_mapper.get_value(row_dict, "验收等级"),
            "project_description": self._column_mapper.get_value(row_dict, "项目简介"),
            "department": department
        }

    def _validate(self, projects: List[Dict[str, Any]]) -> tuple:
        """
        验证项目数据

        Returns:
            (validated_projects, ValidationResult)
        """
        validated = []
        all_content_issues = []
        all_completeness_issues = []

        for idx, project in enumerate(projects):
            content_issues = []
            completeness_issues = []

            # ========== 必需字段检查 ==========
            # 1. 项目名称（必须非空）
            project_name = project.get("project_name", "")
            if not project_name or not str(project_name).strip():
                completeness_issues.append(ValidationError(
                    field_name="project_name",
                    error_type="missing",
                    error_message="缺少项目名称",
                    error_category="completeness"
                ))

            # 2. 年份（必须存在）
            year = project.get("year")
            if year is None:
                completeness_issues.append(ValidationError(
                    field_name="year",
                    error_type="missing",
                    error_message="缺少年份",
                    error_category="completeness"
                ))

            # 3. 项目级别（必须存在）
            project_level = project.get("project_level", "")
            if not project_level or not str(project_level).strip():
                completeness_issues.append(ValidationError(
                    field_name="project_level",
                    error_type="missing",
                    error_message="缺少项目级别",
                    error_category="completeness"
                ))

            # 4. 负责人（姓名和学号）
            leader_name = project.get("leader_name", "")
            leader_sid = project.get("leader_student_id", "")
            if not leader_name or not str(leader_name).strip():
                completeness_issues.append(ValidationError(
                    field_name="leader_name",
                    error_type="missing",
                    error_message="缺少负责人姓名",
                    error_category="completeness"
                ))
            else:
                # 验证姓名格式
                if not DataParser.validate_name(leader_name):
                    content_issues.append(ValidationError(
                        field_name="leader_name",
                        error_type="invalid",
                        error_message=f"负责人姓名格式不正确: {leader_name}",
                        error_category="content",
                        invalid_value=leader_name
                    ))

            if not leader_sid or not str(leader_sid).strip():
                completeness_issues.append(ValidationError(
                    field_name="leader_student_id",
                    error_type="missing",
                    error_message="缺少负责人学号",
                    error_category="completeness"
                ))
            else:
                # 验证学号格式（9位数字）
                if not DataParser.validate_student_id(leader_sid):
                    content_issues.append(ValidationError(
                        field_name="leader_student_id",
                        error_type="invalid",
                        error_message=f"负责人学号格式不正确（应为9位数字）: {leader_sid}",
                        error_category="content",
                        invalid_value=leader_sid
                    ))

            # 5. 指导教师（至少一个）
            supervisors = project.get("supervisors", [])
            if not supervisors or len(supervisors) == 0:
                completeness_issues.append(ValidationError(
                    field_name="supervisors",
                    error_type="missing",
                    error_message="缺少指导教师",
                    error_category="completeness"
                ))

            # ========== 格式验证（内容问题）==========

            # 验证成员学号（从字符串 "姓名(学号)" 中提取学号）
            members = project.get("members", [])
            for member_str in members:
                # 从 "姓名(学号)" 格式中提取学号
                match = re.search(r'\((\d{9})\)', member_str)
                if match:
                    sid = match.group(1)
                    if not DataParser.validate_student_id(sid):
                        content_issues.append(ValidationError(
                            field_name="members",
                            error_type="invalid",
                            error_message=f"成员学号格式不正确: {sid}",
                            error_category="content",
                            invalid_value=member_str
                        ))

            # 记录问题到总列表
            all_content_issues.extend(content_issues)
            all_completeness_issues.extend(completeness_issues)

            # 记录警告日志
            for issue in content_issues:
                logger.warning(f"项目 {idx + 1} 内容问题: {issue.field_name} - {issue.error_message}")
            for issue in completeness_issues:
                logger.warning(f"项目 {idx + 1} 完整性问题: {issue.field_name} - {issue.error_message}")

            validated.append(project)

        # 生成 ValidationResult
        is_valid = len(all_content_issues) == 0 and len(all_completeness_issues) == 0
        validation_result = ValidationResult(
            is_valid=is_valid,
            content_issues=all_content_issues,
            completeness_issues=all_completeness_issues
        )

        return validated, validation_result

    def _other_result(self, message: str) -> ExtractResult:
        """返回other类型结果"""
        return ExtractResult(
            status=ExtractStatus.SUCCESS,
            data={"note": message},
            template_type=TemplateType.OTHER,
            extractor_name=self.name,
        )

    @classmethod
    def from_config_loader(cls, config_loader) -> "InnovationExtractor":
        """从配置加载器创建抽取器"""
        config = config_loader.load_config()
        innovation_cfg = config.get("extract", {}).get("innovation", {})
        return cls(innovation_cfg)
