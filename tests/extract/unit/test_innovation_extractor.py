import pytest

"""
大创抽取器单元测试

测试大创抽取器的各个功能模块。
"""
import pytest
import tempfile
import os
import sys
from pathlib import Path
from openpyxl import Workbook

# 添加项目根到路径
project_root = Path(__file__).resolve().parents[3]
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from backend.extract.extractors import InnovationExtractor
from backend.extract.extractors.innovation import ColumnMapper, DataParser
from backend.extract.extractors.base import ExtractContext
from backend.extract.types import ExtractStatus, TemplateType


# ==================== ColumnMapper 测试 ====================

class TestColumnMapper:
    """列名映射器测试"""

    @pytest.fixture
    def mapper(self):
        """创建列名映射器"""
        mapping = {
            "项目名称": ["项目名称", "名称", "项目"],
            "学号": ["学号", "学生学号", "ID"]
        }
        return ColumnMapper(mapping)

    def test_normalize(self, mapper):
        """测试列名清洗"""
        assert mapper.normalize("  项目名称  ") == "项目名称"
        assert mapper.normalize("项目\n名称") == "项目名称"
        assert mapper.normalize("项目\t名称") == "项目名称"
        assert mapper.normalize(None) == ""

    def test_find_column(self, mapper):
        """测试查找列名"""
        columns = ["系别", "项目名称", "姓名"]
        assert mapper.find_column(columns, "项目名称") == "项目名称"
        assert mapper.find_column(columns, "项目编号") is None

    def test_find_column_with_variants(self, mapper):
        """测试查找列名变体"""
        columns = ["系别", "名称", "姓名"]
        assert mapper.find_column(columns, "项目名称") == "名称"

        columns = ["系别", "项目", "姓名"]
        assert mapper.find_column(columns, "项目名称") == "项目"

    def test_get_value(self, mapper):
        """测试获取值"""
        row_dict = {
            "系别": "计算机工程系",
            "名称": "测试项目",
            "姓名": "张三"
        }
        assert mapper.get_value(row_dict, "项目名称") == "测试项目"
        assert mapper.get_value(row_dict, "学号") is None


# ==================== DataParser 测试 ====================

class TestDataParser:
    """数据解析器测试"""

    def test_parse_teachers(self):
        """测试解析指导教师"""
        assert DataParser.parse_teachers("张老师,李老师") == ["张老师", "李老师"]
        assert DataParser.parse_teachers("张老师、李老师、王老师") == ["张老师", "李老师", "王老师"]
        assert DataParser.parse_teachers("张老师；李老师") == ["张老师", "李老师"]
        assert DataParser.parse_teachers("") == []
        assert DataParser.parse_teachers(None) == []

    def test_parse_students(self):
        """测试解析学生成员"""
        # 姓名/学号格式
        result = DataParser.parse_students("张三/212203227")
        assert result == [{"姓名": "张三", "学号": "212203227"}]

        # 姓名（学号）格式
        result = DataParser.parse_students("张三（212203227）")
        assert result == [{"姓名": "张三", "学号": "212203227"}]

        # 学号+姓名格式
        result = DataParser.parse_students("212203227张三")
        assert result == [{"姓名": "张三", "学号": "212203227"}]

    def test_parse_students_fullwidth_parens_regression(self):
        """T49/T75 挂账修复回归：全角括号无分隔符串不再被兜底正则污染。"""
        # 无分隔符连续串（原缺陷触发形态）
        result = DataParser.parse_students("张三（212203227）李四（212203228）")
        assert result == [
            {"姓名": "张三", "学号": "212203227"},
            {"姓名": "李四", "学号": "212203228"},
        ]

        # 半角括号形态（防修复只顾全角）
        result = DataParser.parse_students("张三(212203227)")
        assert result == [{"姓名": "张三", "学号": "212203227"}]

        # 单个全角括号串（原缺陷最小复现）
        result = DataParser.parse_students("张三（212203227）")
        assert result == [{"姓名": "张三", "学号": "212203227"}]

        # 多个学生逗号分隔
        result = DataParser.parse_students("张三/212203227,李四/212203228")
        assert len(result) == 2
        assert result[0] == {"姓名": "张三", "学号": "212203227"}

    def test_parse_date_range(self):
        """测试解析日期范围"""
        assert DataParser.parse_date_range("2024.6-2025.6") == ("2024.06", "2025.06")
        assert DataParser.parse_date_range("2024.01-2025.12") == ("2024.01", "2025.12")
        # T75 分诊：新增全角～区间解析（改进）
        assert DataParser.parse_date_range("2024～2025") == ("2024.01", "2025.12")
        assert DataParser.parse_date_range(None) == (None, None)
        assert DataParser.parse_date_range("") == (None, None)

    def test_parse_year_to_date_range(self):
        """测试年份转日期范围"""
        assert DataParser.parse_year_to_date_range(2024) == ("2024.06", "2025.06")
        assert DataParser.parse_year_to_date_range("2023") == ("2023.06", "2024.06")
        assert DataParser.parse_year_to_date_range(None) == (None, None)

    def test_validate_student_id(self):
        """测试学号验证（固定9位）"""
        assert DataParser.validate_student_id("212203227") == True
        assert DataParser.validate_student_id("21220322") == False  # 8位
        assert DataParser.validate_student_id("2122032277") == False  # 10位
        assert DataParser.validate_student_id("abcd12345") == False  # 包含字母
        assert DataParser.validate_student_id("") == False

    def test_validate_name(self):
        """测试姓名验证"""
        assert DataParser.validate_name("张三") == True
        assert DataParser.validate_name("李小明") == True
        assert DataParser.validate_name("张") == False  # 太短
        assert DataParser.validate_name("12345") == False  # 纯数字
        assert DataParser.validate_name("张3") == False  # 包含数字


# ==================== InnovationExtractor 测试 ====================

class TestInnovationExtractor:
    """大创抽取器测试"""

    @pytest.fixture
    def config(self):
        """创建配置"""
        return {
            "enabled": True,
            "extensions": [".xlsx", ".xls"],
            "target_departments": ["计算机工程系"],
            "header_keywords": ["项目", "学号", "教师", "负责人"],
            "file_keywords": [
                "大学生创新创业训练计划项目",
                "大学生创新创业",
                "大创"
            ],
            "column_mapping": {
                "项目编号": ["项目编号", "编号"],
                "项目名称": ["项目名称", "名称", "项目"],
                "学生负责人": ["学生负责人", "负责人"],
                "项目负责人学号": ["项目负责人学号", "负责人学号"],
                "项目其他成员信息": ["项目其他成员信息", "成员"],
                "指导教师": ["指导教师", "教师"],
                "项目级别": ["项目级别", "级别"],
            },
            "student_id_length": 9
        }

    @pytest.fixture
    def extractor(self, config):
        """创建抽取器实例"""
        return InnovationExtractor(config)

    def test_init(self, extractor):
        """测试初始化"""
        assert extractor.name == "innovation"
        # T75 分诊：description 类属性文案已精简
        assert extractor.description == "大学生创新创业训练项目文件"
        assert ".xlsx" in extractor.extensions
        assert ".xls" in extractor.extensions

    def test_matches_extension(self, extractor):
        """测试扩展名匹配"""
        assert extractor.matches_extension(".xlsx") == True
        assert extractor.matches_extension(".XLSX") == True
        assert extractor.matches_extension(".xls") == True
        assert extractor.matches_extension(".pdf") == False
        assert extractor.matches_extension(".jpg") == False

    def test_matches_keywords(self, extractor):
        """测试关键词匹配（大创不使用）"""
        assert extractor.matches_keywords("任意文本") == False

    def test_read_xlsx(self, extractor, tmp_path):
        """测试读取xlsx文件"""
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        ws.append(["2025年学院推荐申报国家级、省级大学生创新创业训练计划项目32项"])
        ws.append(["系别", "项目编号", "项目名称", "学生负责人", "项目负责人学号"])
        ws.append(["计算机工程系", "2025001", "测试项目", "张三", "212203227"])
        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        rows = extractor._read_excel(str(file_path))
        assert len(rows) == 3
        assert "大学生创新创业" in rows[0][0]

    def test_is_innovation_file(self, extractor):
        """测试大创文件检测"""
        rows = [
            ["2025年学院推荐申报国家级、省级大学生创新创业训练计划项目32项"],
            ["系别", "项目编号", "项目名称"]
        ]

        # 通过第一行识别
        assert extractor._is_innovation_file(rows, "test.xlsx") == True

        # 通过文件名识别
        assert extractor._is_innovation_file([], "大创项目.xlsx") == True
        assert extractor._is_innovation_file([], "大学生创新创业训练.xlsx") == True

        # 不是大创文件
        assert extractor._is_innovation_file([], "普通文件.xlsx") == False

    def test_find_header_row(self, extractor):
        """测试查找表头行"""
        rows = [
            ["2025年学院推荐申报..."],
            ["系别", "项目编号", "项目名称", "学生负责人"],
            ["数据", "...", "...", "..."]
        ]

        idx = extractor._find_header_row(rows)
        assert idx == 1

    def test_extract_level_from_first_row(self, extractor):
        """测试从第一行提取级别"""
        assert extractor._extract_level_from_first_row("2025年国家级大学生创新创业") == "国家级"
        assert extractor._extract_level_from_first_row("2023年院级大学生创新创业") == "院级"
        assert extractor._extract_level_from_first_row("普通文本") == ""

    def test_extract_year_from_first_row(self, extractor):
        """测试从第一行提取年份"""
        assert extractor._extract_year_from_first_row("2025年学院推荐申报") == 2025
        assert extractor._extract_year_from_first_row("2023年院级项目") == 2023
        assert extractor._extract_year_from_first_row("普通文本") == None

    def test_extract_with_mock_context(self, extractor, tmp_path):
        """测试使用Mock上下文进行抽取"""
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        ws.append(["2025年国家级大学生创新创业训练计划项目"])
        ws.append(["系别", "项目编号", "项目名称", "学生负责人", "项目负责人学号", "项目其他成员信息", "指导教师"])
        ws.append(["计算机工程系", "2025001", "测试项目A", "张三", "212203227", "李四/212203228", "王老师"])
        ws.append(["数学系", "2025002", "测试项目B", "李四", "212203228", "王五/212203229", "赵老师"])
        file_path = tmp_path / "test_innovation.xlsx"
        wb.save(str(file_path))

        # 创建抽取上下文
        ctx = ExtractContext(
            file_path=str(file_path),
            ocr_text=None,
            use_ocr_cache=False,
            use_llm_cache=False,
        )

        # 执行抽取
        result = extractor.extract(ctx)

        # 验证结果
        assert result.status == ExtractStatus.SUCCESS
        assert result.template_type == TemplateType.INNOVATION
        assert result.extractor_name == "innovation"
        assert "projects" in result.data
        assert result.data["count"] == 1  # 只有计算机工程系的

        projects = result.data["projects"]
        assert len(projects) == 1
        # T75 分诊：输出键名约定中改英
        assert projects[0]["project_name"] == "测试项目A"
        assert projects[0]["department"] == "计算机工程系"

    def test_extract_non_innovation_file(self, extractor, tmp_path):
        """测试抽取非大创文件"""
        # 创建普通Excel文件
        wb = Workbook()
        ws = wb.active
        ws.append(["普通文件标题"])
        ws.append(["姓名", "年龄"])
        ws.append(["张三", "20"])
        file_path = tmp_path / "test_normal.xlsx"
        wb.save(str(file_path))

        ctx = ExtractContext(
            file_path=str(file_path),
            ocr_text=None,
            use_ocr_cache=False,
            use_llm_cache=False,
        )

        result = extractor.extract(ctx)

        # 应该返回other
        assert result.template_type == TemplateType.OTHER
        assert "不是大创文件" in result.data.get("note", "")

    def test_extract_with_no_data(self, extractor, tmp_path):
        """测试抽取大创文件但没有目标系别的数据"""
        # 创建只有其他系别的Excel文件
        wb = Workbook()
        ws = wb.active
        ws.append(["2025年国家级大学生创新创业训练计划项目"])
        ws.append(["系别", "项目编号", "项目名称", "学生负责人"])
        ws.append(["数学系", "2025001", "测试项目A", "张三"])
        file_path = tmp_path / "test_no_target.xlsx"
        wb.save(str(file_path))

        ctx = ExtractContext(
            file_path=str(file_path),
            ocr_text=None,
            use_ocr_cache=False,
            use_llm_cache=False,
        )

        result = extractor.extract(ctx)

        # 应该返回other
        assert result.template_type == TemplateType.OTHER


# ==================== 集成测试辅助函数 ====================

def create_test_innovation_xlsx(path: Path, include_target_dept: bool = True):
    """
    创建测试用的大创Excel文件

    Args:
        path: 文件保存路径
        include_target_dept: 是否包含目标系别的数据
    """
    wb = Workbook()
    ws = wb.active

    # 第一行：标题
    ws.append(["2025年国家级大学生创新创业训练计划项目"])

    # 第二行：表头
    ws.append(["系别", "项目编号", "项目名称", "学生负责人", "项目负责人学号", "项目其他成员信息", "指导教师"])

    # 添加计算机工程系的数据
    if include_target_dept:
        ws.append(["计算机工程系", "2025001", "基于深度学习的交通事故可视化分析", "袁凯凯", "212203227", "宋家豪/212203264,李佳豪/212203250", "胡清桂"])
        ws.append(["计算机工程系", "2025002", "智能交通管理系统", "张三", "212203228", "李四/212203229", "王老师"])

    # 添加其他系的数据
    ws.append(["数学系", "2025003", "数学建模研究", "王五", "212203230", "赵六/212203231", "孙老师"])

    wb.save(str(path))


def create_normal_xlsx(path: Path):
    """创建普通的非大创Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.append(["学生信息表"])
    ws.append(["姓名", "年龄", "班级"])
    ws.append(["张三", "20", "计科1班"])
    wb.save(str(path))


if __name__ == "__main__":
    # 运行测试
    pytest.main([__file__, "-v"])

